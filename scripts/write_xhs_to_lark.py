#!/usr/bin/env python3
"""
小红书数据写入飞书多维表格 — 内置于 skill
写入两张表：账号数据(每日) + 内容数据(笔记)
不写周数据表（用户暂不需要）

数据源：小红书创作者中心导出的5个Excel文件
依赖：openpyxl, lark-cli
Table ID 配置：config/table_mapping.json（自动加载，无需命令行传入）

用法：
  python3 write_xhs_to_lark.py --account 火车票小红书 --dir <导出目录> \
    --friday 2026-08-14 --thursday 2026-08-20

harness 加固：
  1. validate_account_login — 写入前 whoami 校验当前登录账号
  2. validate_unique — 检查日报数据是否全0
  3. check_and_cleanup_duplicates — 写入后按主键去重
"""

import openpyxl
import json
import subprocess
import os
import re
import argparse


def run_lark(args, json_body=None, cwd='/tmp'):
    """执行 lark-cli 命令，--json 用 @file 相对路径传参"""
    if json_body is not None:
        fname = 'lark_payload.json'
        with open(os.path.join(cwd, fname), 'w', encoding='utf-8') as f:
            json.dump(json_body, f, ensure_ascii=False)
        args = args + ['--json', f'@{fname}']
    result = subprocess.run(
        ['lark-cli'] + args,
        capture_output=True, text=True, cwd=cwd
    )
    try:
        return json.loads(result.stdout)
    except:
        print(f"  ❌ lark-cli 返回非JSON: stdout={result.stdout[:300]}, stderr={result.stderr[:300]}")
        return None


def parse_trend(wb, sheet_name):
    """解析趋势sheet，返回 {日期: 数值}。
    支持带单位的字符串如 "8%"→0.08, "18秒"→18
    """
    if sheet_name not in wb.sheetnames:
        return {}
    sh = wb[sheet_name]
    data = {}
    date_re = re.compile(r'(\d{4})年(\d{2})月(\d{2})日')
    for r in range(2, sh.max_row + 1):
        raw_date = sh.cell(r, 1).value
        raw_val = sh.cell(r, 2).value
        if not raw_date:
            continue
        if isinstance(raw_date, str):
            m = date_re.search(raw_date)
            if m:
                date_str = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
            else:
                continue
        elif hasattr(raw_date, 'strftime'):
            date_str = raw_date.strftime('%Y-%m-%d')
        else:
            continue

        if raw_val is None:
            val = 0
        elif isinstance(raw_val, (int, float)):
            val = raw_val
        elif isinstance(raw_val, str):
            num_match = re.search(r'[\d.]+', raw_val)
            if num_match:
                num = float(num_match.group())
                val = num / 100 if '%' in raw_val else num
            else:
                val = 0
        else:
            val = 0
        data[date_str] = val
    return data


def parse_summary(wb, sheet_name=None):
    """解析汇总sheet（指标-值），返回 {指标名: 数值}"""
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    sh = wb[sheet_name]
    data = {}
    for r in range(2, sh.max_row + 1):
        key = sh.cell(r, 1).value
        val = sh.cell(r, 2).value
        if key and val is not None:
            if isinstance(val, str):
                num_match = re.search(r'[\d.]+', val)
                if num_match:
                    val = float(num_match.group())
                    if '%' in val:
                        val = val / 100
            data[key] = val
    return data


def parse_content_data(path):
    """读取内容分析笔记明细"""
    wb = openpyxl.load_workbook(path)
    sh = wb[wb.sheetnames[0]]
    records = []
    for r in range(3, sh.max_row + 1):
        title = sh.cell(r, 1).value
        if not title:
            continue
        publish_time = sh.cell(r, 2).value
        tixing = sh.cell(r, 3).value
        exposure = sh.cell(r, 4).value or 0
        views = sh.cell(r, 5).value or 0
        cover_ctr = sh.cell(r, 6).value or 0
        likes = sh.cell(r, 7).value or 0
        comments = sh.cell(r, 8).value or 0
        favorites = sh.cell(r, 9).value or 0
        new_followers = sh.cell(r, 10).value or 0
        shares = sh.cell(r, 11).value or 0
        avg_watch = sh.cell(r, 12).value or 0
        danmu = sh.cell(r, 13).value or 0
        total_interact = int(likes + comments + favorites + shares)

        if isinstance(publish_time, str):
            pt = publish_time.replace('年','-').replace('月','-').replace('日','T') \
                .replace('时',':').replace('分',':').replace('秒','')+'+08:00'
        else:
            pt = publish_time

        records.append({
            '笔记标题': title,
            '首次发布时间': pt,
            '体裁': tixing,
            '曝光': int(exposure),
            '观看量': int(views),
            '封面点击率': round(cover_ctr, 4),
            '点赞': int(likes),
            '收藏': int(favorites),
            '评论': int(comments),
            '分享': int(shares),
            '涨粉': int(new_followers),
            '人均观看时长（s）': int(avg_watch),
            '弹幕': int(danmu),
            '总互动': total_interact,
        })
    return records


# ===== harness: 数据安全校验 =====

def validate_account_login(account_name):
    """写入前校验当前 opencli 登录的小红书账号是否匹配。
    匹配规则：whoami 返回的 username 包含 account name 核心关键词。
    """
    result = subprocess.run(
        ['opencli', 'xiaohongshu', 'whoami'],
        capture_output=True, text=True
    )
    output = result.stdout
    if 'logged_in: true' not in output:
        print("  ❌ 未登录")
        return False

    username = ''
    for line in output.split('\n'):
        if 'username:' in line.lower():
            username = line.split('username:')[-1].strip()
            break
    print(f"  ℹ️ 当前登录: {username}")

    if not username:
        print("  ⚠️ 无法确认用户名，放行但需人工核实")
        return True

    keywords = []
    if '火车票' in account_name:
        keywords.append('火车票')
    if '旅行' in account_name:
        keywords.append('旅行')
    if '员工' in account_name:
        keywords.append('员工')

    if any(kw in username for kw in keywords):
        return True
    else:
        print(f"  🚨 账号不匹配！期望关键词: {keywords}, 实际: {username}")
        return False


def validate_unique(daily_records):
    """检查日报数据是否全0（可能导出失败的信号）"""
    if daily_records:
        first = daily_records[0]
        all_zero = all(v == 0 for k, v in first.items() if k != '日期')
        if all_zero:
            print("  🚨 日报数据全0，可能导出失败")
            return False
    return True


def check_and_cleanup_duplicates(base_token, table_id, search_field, search_keyword, dedup_key_fields):
    """写入后去重检查：搜索匹配记录，按 dedup_key_fields 判断重复，保留最早写入的一条。
    """
    result = subprocess.run([
        'lark-cli', 'base', '+record-search',
        '--base-token', base_token,
        '--table-id', table_id,
        '--as', 'user',
        '--format', 'json',
        '--keyword', search_keyword,
        '--search-field', search_field,
        '--limit', '50'
    ], capture_output=True, text=True)

    try:
        out = json.loads(result.stdout)
    except:
        return 0

    if not out.get('ok') or not out['data'].get('data'):
        return 0

    rows = out['data']['data']
    fields = out['data']['fields']
    record_ids = out['data']['record_id_list']

    # 找到 dedup_key_fields 的列索引
    key_indices = []
    for dkf in dedup_key_fields:
        for i, fname in enumerate(fields):
            if fname == dkf:
                key_indices.append(i)
                break

    # 按主键值分组
    groups = {}
    for i, row in enumerate(rows):
        key = tuple(str(row[idx]) for idx in key_indices)
        groups.setdefault(key, []).append(record_ids[i])

    # 找出重复组，保留第一个，删除其余
    dup_ids = []
    for key, rids in groups.items():
        if len(rids) > 1:
            dup_ids.extend(rids[1:])

    if dup_ids:
        print(f"  🚨 发现 {len(dup_ids)} 条重复记录，正在清理...")
        del_body = {"record_id_list": dup_ids}
        run_lark([
            'base', '+record-delete',
            '--base-token', base_token,
            '--table-id', table_id,
            '--as', 'user', '--yes'
        ], json_body=del_body)
        print(f"  ✅ 已删除 {len(dup_ids)} 条重复记录")
    else:
        print(f"  ✅ 无重复记录（{len(rows)} 条已检查）")
    return len(dup_ids)


def delete_existing_records(base_token, table_id, keyword, search_field='日期'):
    """删除已存在的匹配记录"""
    result = subprocess.run([
        'lark-cli', 'base', '+record-search',
        '--base-token', base_token,
        '--table-id', table_id,
        '--as', 'user',
        '--format', 'json',
        '--keyword', keyword,
        '--search-field', search_field,
        '--limit', '50'
    ], capture_output=True, text=True)

    try:
        out = json.loads(result.stdout)
    except:
        return 0

    if out.get('ok') and out['data'].get('record_id_list'):
        ids = out['data']['record_id_list']
        del_body = {"record_id_list": ids}
        run_lark([
            'base', '+record-delete',
            '--base-token', base_token,
            '--table-id', table_id,
            '--as', 'user', '--yes'
        ], json_body=del_body)
        print(f"  删除 {len(ids)} 条旧记录")
        return len(ids)
    return 0


def batch_create(base_token, table_id, records):
    if not records:
        print("  ⚠️ 无记录可写入")
        return False
    body = {"create_records": records}
    out = run_lark([
        'base', '+record-batch-create',
        '--base-token', base_token,
        '--table-id', table_id,
        '--as', 'user'
    ], json_body=body)

    if out and out.get('ok'):
        print(f"  ✅ 写入 {len(records)} 条记录")
        return True
    else:
        print(f"  ❌ 写入失败: {out.get('error') if out else 'No response'}")
        return False


def main():
    parser = argparse.ArgumentParser(description='小红书数据写入飞书多维表格')
    parser.add_argument('--account', required=True, help='账号名称（火车票小红书/旅行小红书/员工号）')
    parser.add_argument('--dir', required=True, help='导出目录路径')
    parser.add_argument('--friday', required=True, help='周五日期 YYYY-MM-DD')
    parser.add_argument('--thursday', required=True, help='周四日期 YYYY-MM-DD')
    args = parser.parse_args()
    
    # 从配置文件加载 table ID
    try:
        config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 
                                   'config', 'table_mapping.json')
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        if args.account not in config:
            print(f"❌ 账号 {args.account} 未在配置文件中找到")
            return
        
        table_config = config[args.account]
        base_token = table_config['base_token']
        daily_table = table_config['daily_table']
        content_table = table_config['content_table']
        print(f"✅ 已加载配置：base_token={base_token[:20]}...")
    except Exception as e:
        print(f"❌ 加载配置失败：{e}")
        return
    
    d = args.dir

    print(f"\n{'='*60}")
    print(f"处理账号: {args.account}")
    print(f"{'='*60}")

    # === 前置校验 ===
    print("\n🔍 [1] 验证数据来源...")
    if not validate_account_login(args.account):
        print("  ❌ 账号验证失败，跳过")
        return

    # 读取4个汇总Excel
    watch_wb = openpyxl.load_workbook(os.path.join(d, '观看数据.xlsx'))
    interact_wb = openpyxl.load_workbook(os.path.join(d, '互动数据.xlsx'))
    fans_wb = openpyxl.load_workbook(os.path.join(d, '涨粉数据.xlsx'))
    publish_wb = openpyxl.load_workbook(os.path.join(d, '发布数据.xlsx'))

    # === 解析每日趋势数据 ===
    print("\n📊 [2] 解析每日趋势数据...")
    exposure = parse_trend(watch_wb, '曝光趋势')
    views = parse_trend(watch_wb, '观看趋势')
    cover_ctr = parse_trend(watch_wb, '封面点击率趋势')
    avg_watch = parse_trend(watch_wb, '平均观看时长趋势')
    total_watch = parse_trend(watch_wb, '观看总时长趋势')
    completion = parse_trend(watch_wb, '视频完播率趋势')

    likes = parse_trend(interact_wb, '点赞趋势')
    comments = parse_trend(interact_wb, '评论趋势')
    favorites = parse_trend(interact_wb, '收藏趋势')
    shares = parse_trend(interact_wb, '分享趋势')

    net_fans = parse_trend(fans_wb, '净涨粉趋势')
    new_follows = parse_trend(fans_wb, '新增关注趋势')
    cancels = parse_trend(fans_wb, '取消关注趋势')
    profile_visits = parse_trend(fans_wb, '主页访客趋势')
    profile_rate = parse_trend(fans_wb, '主页转粉率趋势')

    total_pub = parse_trend(publish_wb, '总发布趋势')
    video_pub = parse_trend(publish_wb, '发布视频趋势')
    image_pub = parse_trend(publish_wb, '发布图文趋势')

    dates = sorted(set(exposure.keys()))
    print(f"  日期范围: {dates[0]} 至 {dates[-1]}, 共{len(dates)}天")

    daily_records = []
    for day in dates:
        day_iso = f"{day}T00:00:00+08:00"
        daily_records.append({
            '日期': day_iso,
            '曝光': int(exposure.get(day, 0)),
            '观看': int(views.get(day, 0)),
            '封面点击率': round(cover_ctr.get(day, 0), 4),
            '平均观看时长（s）': round(avg_watch.get(day, 0), 3),
            '观看总时长（s）': int(total_watch.get(day, 0)),
            '视频完播率': round(completion.get(day, 0), 4),
            '点赞': int(likes.get(day, 0)),
            '评论': int(comments.get(day, 0)),
            '收藏': int(favorites.get(day, 0)),
            '分享': int(shares.get(day, 0)),
            '新增关注': int(new_follows.get(day, 0)),
            '取消关注': int(cancels.get(day, 0)),
            '净涨粉': int(net_fans.get(day, 0)),
            '主页访客': int(profile_visits.get(day, 0)),
            '主页转粉率': round(profile_rate.get(day, 0), 4),
            '发布视频条数': int(video_pub.get(day, 0)),
            '发布图文条数': int(image_pub.get(day, 0)),
            '总发布条数': int(total_pub.get(day, 0)),
        })

    # === 数据校验 ===
    print("\n🔍 [3] 数据校验...")
    if not validate_unique(daily_records):
        print("  ❌ 数据校验失败")
        return
    for r in daily_records:
        print(f"  {r['日期'][:10]}: 曝光={r['曝光']}, 观看={r['观看']}, 涨粉={r['净涨粉']}")

    # === 写入账号数据(每日) ===
    print(f"\n📝 [4] 写入账号数据表(每日)...")
    # 前置删除本周已有数据
    for day in dates:
        delete_existing_records(base_token, daily_table, day.replace('-', '/'), '日期')
    batch_create(base_token, daily_table, daily_records)
    # 写入后去重校验
    check_and_cleanup_duplicates(base_token, daily_table, '日期',
                                 dates[0].replace('-', '/'), ['日期'])

    # === 写入内容数据(笔记) ===
    print(f"\n📝 [5] 写入内容数据表(笔记)...")
    content_path = os.path.join(d, '内容分析_笔记明细.xlsx')
    if os.path.exists(content_path):
        notes = parse_content_data(content_path)
        print(f"  共 {len(notes)} 条笔记")
        for n in notes:
            print(f"    - {n['笔记标题'][:30]}... 曝光={n['曝光']}, 点赞={n['点赞']}")
        # 前置去重：按每篇笔记标题搜索，已存在则先删
        for n in notes:
            check_and_cleanup_duplicates(base_token, content_table, '笔记标题',
                                         n['笔记标题'][:10], ['笔记标题', '首次发布时间'])
        batch_create(base_token, content_table, notes)
        # 写入后去重校验
        for n in notes:
            check_and_cleanup_duplicates(base_token, content_table, '笔记标题',
                                         n['笔记标题'][:10], ['笔记标题', '首次发布时间'])
    else:
        print("  ⚠️ 内容分析文件不存在")

    print(f"\n✅ {args.account} 全部写入完成")


if __name__ == '__main__':
    main()
